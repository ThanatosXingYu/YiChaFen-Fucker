from __future__ import annotations

import csv
import json
import os
import queue
import re
import shutil
import threading
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side

from .config import DEFAULT_REFRESH_CLASS, DEFAULT_REFRESH_NAME, AppPaths
from .models import QueryAttempt, QueryFailure, QueryField, QueryInfo
from .utils import clean_text


class YichafenParser:
    post_pattern = re.compile(r"\$\.post\(\s*['\"]([^'\"]+)['\"]", re.I)
    lock_post_pattern = re.compile(
        r"doLockBtn[\s\S]{0,3000}?\$\.post\(\s*['\"]([^'\"]+)['\"]",
        re.I,
    )
    lock_success_pattern = re.compile(
        r"doLockBtn[\s\S]{0,3000}?location\.href\s*=\s*['\"]([^'\"]+)['\"]",
        re.I,
    )

    @classmethod
    def parse_home(cls, page_url: str, html: str) -> list[QueryInfo]:
        soup = BeautifulSoup(html, "html.parser")
        queries: list[QueryInfo] = []
        seen: set[str] = set()
        for anchor in soup.select("a[href*='/qz/']"):
            href = anchor.get("href", "").strip()
            if not href:
                continue
            url = urljoin(page_url, href)
            if url in seen:
                continue
            seen.add(url)

            title_node = anchor.select_one("p")
            name = title_node.get_text(" ", strip=True) if title_node else ""
            if not name:
                name = anchor.get_text(" ", strip=True) or url
            queries.append(QueryInfo(name=name, url=url))
        return queries

    @classmethod
    def parse_query_page(cls, page_url: str, html: str) -> tuple[str, list[QueryField], str]:
        soup = BeautifulSoup(html, "html.parser")
        title_node = soup.select_one(".s_page-header_color, .s_page-header_pic .banner-text")
        title = title_node.get_text(" ", strip=True) if title_node else ""
        if not title:
            title_tag = soup.find("title")
            title = title_tag.get_text(" ", strip=True) if title_tag else page_url

        form = soup.select_one("#form-data") or soup.find("form")
        if form is None:
            raise ValueError("查询页没有找到 form-data 表单")

        fields: list[QueryField] = []
        seen_names: set[str] = set()
        for control in form.select("input[name], select[name], textarea[name]"):
            name = control.get("name", "").strip()
            if not name or name in seen_names:
                continue
            control_type = control.get("type", "").lower()
            if control_type in {"hidden", "button", "submit", "image", "reset"}:
                continue
            if name in {"verify", "NECaptchaValidate"}:
                continue

            label = control.get("data-sname", "").strip()
            if not label:
                cell = control.find_parent(class_="weui-cell")
                label_node = cell.select_one("label") if cell else None
                label = label_node.get_text(" ", strip=True) if label_node else ""
            if not label:
                label = control.get("placeholder", "").replace("请输入", "").strip()
            if not label:
                label = name

            seen_names.add(name)
            fields.append(QueryField(label=label, name=name))

        if not fields:
            raise ValueError("查询页没有解析到任何查询条件字段")

        match = cls.post_pattern.search(html)
        if not match:
            raise ValueError("查询页没有解析到 $.post 提交接口")
        post_url = urljoin(page_url, match.group(1))
        return title, fields, post_url

    @classmethod
    def parse_result(cls, html: str) -> dict[str, str]:
        soup = BeautifulSoup(html, "html.parser")
        plain = soup.get_text(" ", strip=True)
        if "来源异常" in plain:
            raise ValueError("结果页提示来源异常，可能缺少 Referer 或 Cookie")

        data: dict[str, str] = {}
        tables = soup.select("#result_content .js_result_table")
        if not tables:
            tables = soup.select("#result_data_table table, .js_result_table, table")

        # 一个结果页可能包含多张结果表，不能横向合并成“姓名_2/班级_2”。
        # 批量查询保存时按“一次查询 -> 一行 CSV”处理，只取当前页第一张结果表。
        if tables:
            table = tables[0]
            for tr in table.select("tr"):
                cells = tr.find_all(["th", "td"])
                if len(cells) < 2:
                    continue
                key = cells[0].get_text(" ", strip=True).rstrip(":：")
                value = cells[1].get_text(" ", strip=True)
                if not key:
                    continue
                if key not in data:
                    data[key] = value

        # 兼容少数证书式结果页：左侧行头 + 右侧值不一定在 table 内。
        if not data:
            for row in soup.select(".cert-row"):
                key_node = row.select_one(".b-txt")
                if not key_node:
                    continue
                key = key_node.get_text(" ", strip=True).rstrip(":：")
                row_text = row.get_text(" ", strip=True)
                value = row_text.replace(key_node.get_text(" ", strip=True), "", 1).strip(" :：")
                if key:
                    if key not in data:
                        data[key] = value

        if not data:
            raise ValueError("结果页没有解析到竖排结果表")
        return data

    @classmethod
    def parse_lock_action(cls, page_url: str, html: str) -> str:
        post_url, _ = cls.parse_lock_request(page_url, html)
        return post_url

    @classmethod
    def parse_lock_request(cls, page_url: str, html: str) -> tuple[str, str]:
        soup = BeautifulSoup(html, "html.parser")
        if not soup.select_one("#doLockBtn"):
            raise ValueError("结果页没有找到锁定按钮")

        lock_script = ""
        for script in soup.find_all("script"):
            text = script.get_text("\n")
            if "doLockBtn" not in text:
                continue
            start = text.find("doLockBtn")
            if start < 0:
                continue
            lock_script = text[start:]
            for marker in ('$("#doSaveEditedFieldsBtn")', "$('#doSaveEditedFieldsBtn')", "function checkData"):
                marker_index = lock_script.find(marker)
                if marker_index > 0:
                    lock_script = lock_script[:marker_index]
                    break
            break

        if not lock_script:
            raise ValueError("结果页没有找到锁定按钮脚本")

        post_match = cls.lock_post_pattern.search(lock_script)
        if not post_match:
            raise ValueError("结果页没有锁定提交接口，当前查询可能未开启锁定")

        success_url = ""
        success_match = cls.lock_success_pattern.search(lock_script)
        if success_match:
            success_url = urljoin(page_url, success_match.group(1).strip())
        return urljoin(page_url, post_match.group(1).strip()), success_url


class YichafenClient:
    def __init__(self, paths: AppPaths, cookie_name: str, timeout: int = 20):
        self.paths = paths
        self.timeout = timeout
        self.session = requests.Session()
        safe_name = re.sub(r"[^a-zA-Z0-9_.-]+", "_", cookie_name)
        self.cookie_file = self.paths.cookies / f"{safe_name}.json"
        self.query_url = ""
        self.base_url = ""
        self.origin = ""
        self.default_headers = {
            "User-Agent": (
                "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
                "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 "
                "Mobile/15E148 Safari/604.1"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.7",
        }
        self._load_cookies()

    def _load_cookies(self) -> None:
        if not self.cookie_file.exists():
            return
        try:
            cookies = json.loads(self.cookie_file.read_text(encoding="utf-8"))
            if isinstance(cookies, dict):
                self.session.cookies.update(cookies)
        except Exception:
            return

    def _save_cookies(self) -> None:
        self.cookie_file.write_text(
            json.dumps(self.session.cookies.get_dict(), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _request(self, method: str, url: str, **kwargs: Any) -> requests.Response:
        headers = dict(self.default_headers)
        headers.update(kwargs.pop("headers", {}) or {})
        response = self.session.request(
            method,
            url,
            headers=headers,
            timeout=self.timeout,
            **kwargs,
        )
        self._save_cookies()
        response.raise_for_status()
        return response

    def open_query_page(self, query_url: str) -> str:
        parsed = urlparse(query_url)
        self.query_url = query_url
        self.base_url = f"{parsed.scheme}://{parsed.netloc}"
        self.origin = self.base_url
        response = self._request("GET", query_url)
        return response.text

    def get_html(self, url: str, referer: str | None = None) -> str:
        headers = {}
        if referer:
            headers["Referer"] = referer
        response = self._request("GET", url, headers=headers)
        return response.text

    def submit_conditions(self, post_url: str, params: dict[str, str]) -> dict[str, Any]:
        if not self.query_url:
            raise RuntimeError("请先打开查询页再提交条件")
        headers = {
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Origin": self.origin,
            "Referer": self.query_url,
            "X-Requested-With": "XMLHttpRequest",
        }
        response = self._request("POST", post_url, headers=headers, data=params)
        try:
            return response.json()
        except ValueError as exc:
            raise ValueError(f"提交接口未返回 JSON：{response.text[:200]}") from exc

    def get_result_page(self, result_url: str) -> str:
        url = urljoin(self.base_url + "/", result_url)
        response = self._request("GET", url, headers={"Referer": self.query_url})
        return response.text

    def submit_lock(self, lock_url: str, referer: str) -> dict[str, Any]:
        headers = {
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Origin": self.origin,
            "Referer": referer,
            "X-Requested-With": "XMLHttpRequest",
        }
        response = self._request("POST", lock_url, headers=headers, data={})
        try:
            return response.json()
        except ValueError as exc:
            raise ValueError(f"锁定接口未返回 JSON：{response.text[:200]}") from exc

    def get_captcha_image(self) -> bytes:
        url = urljoin(self.base_url + "/", f"/public/verify.html?random={time.time()}")
        response = self._request(
            "GET",
            url,
            headers={
                "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
                "Referer": self.query_url,
            },
        )
        return response.content


class CaptchaSolver:
    def __init__(self, paths: AppPaths):
        self.paths = paths
        self._ocr = None
        self._lock = threading.Lock()

    def _ensure_ocr(self) -> Any:
        if self._ocr is not None:
            return self._ocr
        with self._lock:
            if self._ocr is None:
                import ddddocr

                self._ocr = ddddocr.DdddOcr(show_ad=False)
        return self._ocr

    def solve(self, image_bytes: bytes) -> str:
        ocr = self._ensure_ocr()
        with self._lock:
            result = ocr.classification(image_bytes)
        return re.sub(r"\s+", "", str(result or ""))


class DataInputLoader:
    @staticmethod
    def load_delimited(path: Path, delimiter: str = ",") -> pd.DataFrame:
        if path.suffix.lower() not in {".csv", ".txt"}:
            raise ValueError("仅支持 .csv 或 .txt 分隔文本")
        if not delimiter:
            raise ValueError("分隔符不能为空")

        last_error: Exception | None = None
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                dataframe = pd.read_csv(
                    path,
                    dtype=str,
                    keep_default_na=False,
                    encoding=encoding,
                    sep=delimiter,
                    engine="python",
                )
                return DataInputLoader.clean_columns(dataframe)
            except Exception as exc:
                last_error = exc
        raise ValueError(f"CSV 读取失败：{last_error}")

    @staticmethod
    def excel_sheets(path: Path) -> list[str]:
        try:
            workbook = pd.ExcelFile(path)
        except Exception as exc:
            raise ValueError(f"Excel 文件读取失败：{exc}") from exc
        return list(workbook.sheet_names)

    @staticmethod
    def load_excel(path: Path, sheet_name: str | int = 0) -> pd.DataFrame:
        try:
            dataframe = pd.read_excel(path, sheet_name=sheet_name, dtype=str, keep_default_na=False)
        except Exception as exc:
            raise ValueError(f"Excel 工作表解析失败：{exc}") from exc
        return DataInputLoader.clean_columns(dataframe)

    @staticmethod
    def clean_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
        keep_columns = []
        clean_names = []
        seen: dict[str, int] = {}
        for column in dataframe.columns:
            name = str(column).strip()
            if not name or re.fullmatch(r"Unnamed:\s*\d+", name, flags=re.I):
                continue
            count = seen.get(name, 0)
            seen[name] = count + 1
            if count:
                name = f"{name}_{count + 1}"
            keep_columns.append(column)
            clean_names.append(name)
        dataframe = dataframe.loc[:, keep_columns].copy()
        dataframe.columns = clean_names
        return dataframe


# 兼容旧导入名。
CsvInputLoader = DataInputLoader


class ResultCsvWriter:
    def __init__(self, path: Path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.lock = threading.Lock()
        self.headers: list[str] = []
        if self.path.exists() and self.path.stat().st_size > 0:
            with self.path.open("r", encoding="utf-8-sig", newline="") as file:
                reader = csv.DictReader(file)
                self.headers = list(reader.fieldnames or [])

    def append(self, row: dict[str, str]) -> None:
        with self.lock:
            clean_row = {str(key): clean_text(value) for key, value in row.items()}
            if not self.headers:
                self.headers = list(clean_row.keys())
                self._append_row(clean_row, write_header=True)
                return

            extra_headers = [key for key in clean_row if key not in self.headers]
            if extra_headers:
                self._rewrite_with_headers(self.headers + extra_headers)
            self._append_row(clean_row, write_header=False)

    def _append_row(self, row: dict[str, str], write_header: bool) -> None:
        encoding = "utf-8-sig" if write_header else "utf-8"
        with self.path.open("a", encoding=encoding, newline="") as file:
            writer = csv.DictWriter(file, fieldnames=self.headers)
            if write_header:
                writer.writeheader()
            writer.writerow({key: row.get(key, "") for key in self.headers})
            file.flush()
            os.fsync(file.fileno())

    def _rewrite_with_headers(self, new_headers: list[str]) -> None:
        rows: list[dict[str, str]] = []
        if self.path.exists() and self.path.stat().st_size > 0:
            with self.path.open("r", encoding="utf-8-sig", newline="") as file:
                rows = list(csv.DictReader(file))
        self.headers = new_headers
        with self.path.open("w", encoding="utf-8-sig", newline="") as file:
            writer = csv.DictWriter(file, fieldnames=self.headers)
            writer.writeheader()
            for row in rows:
                writer.writerow({key: row.get(key, "") for key in self.headers})
            file.flush()
            os.fsync(file.fileno())


class ResultXlsxWriter:
    def __init__(self, path: Path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.lock = threading.Lock()
        self.headers: list[str] = []
        if self.path.exists() and self.path.stat().st_size > 0:
            self.workbook = load_workbook(self.path)
            self.sheet = self.workbook.active
            self.headers = [
                clean_text(self.sheet.cell(row=1, column=index).value)
                for index in range(1, self.sheet.max_column + 1)
            ]
            self.headers = [header for header in self.headers if header]
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = "查询结果"

        self.alignment = Alignment(horizontal="center", vertical="center")
        side = Side(style="thin", color="000000")
        self.border = Border(left=side, right=side, top=side, bottom=side)

    def append(self, row: dict[str, str]) -> None:
        with self.lock:
            clean_row = {str(key): clean_text(value) for key, value in row.items()}
            if not self.headers:
                self.headers = list(clean_row.keys())
                self.sheet.append(self.headers)
            else:
                extra_headers = [key for key in clean_row if key not in self.headers]
                if extra_headers:
                    self.headers.extend(extra_headers)
                    for column_index, header in enumerate(self.headers, start=1):
                        self.sheet.cell(row=1, column=column_index, value=header)

            self.sheet.append([clean_row.get(header, "") for header in self.headers])
            self._style_used_range()
            self.workbook.save(self.path)

    def _style_used_range(self) -> None:
        for row in self.sheet.iter_rows(
            min_row=1,
            max_row=self.sheet.max_row,
            min_col=1,
            max_col=max(1, len(self.headers)),
        ):
            for cell in row:
                cell.alignment = self.alignment
                cell.border = self.border


def make_result_writer(path: Path) -> ResultCsvWriter | ResultXlsxWriter:
    if path.suffix.lower() == ".xlsx":
        return ResultXlsxWriter(path)
    return ResultCsvWriter(path)


class CacheCleaner:
    @staticmethod
    def clean(paths: AppPaths, *, include_logs: bool = False) -> int:
        targets = [paths.cookies, paths.ocr, paths.temp, paths.tools]
        if include_logs:
            targets.append(paths.logs)

        removed = 0
        for target in targets:
            target.mkdir(parents=True, exist_ok=True)
            for child in target.iterdir():
                if child.name == ".gitkeep":
                    continue
                if child.is_dir():
                    shutil.rmtree(child)
                else:
                    child.unlink()
                removed += 1
        return removed


class QueryRunner:
    def __init__(
        self,
        *,
        paths: AppPaths,
        events: queue.Queue,
        query_info: QueryInfo,
        fields: list[QueryField],
        post_url: str,
        rows: list[dict[str, str]],
        mappings: dict[str, str],
        output_path: Path,
        strategy: str,
        ocr_retries: int,
        refresh_name: str,
        refresh_class: str,
        multithread: bool,
        thread_count: int,
        task_mode: str = "query",
    ):
        self.paths = paths
        self.events = events
        self.query_info = query_info
        self.fields = fields
        self.post_url = post_url
        self.rows = rows
        self.mappings = mappings
        self.output_path = output_path
        self.strategy = strategy
        self.ocr_retries = max(1, ocr_retries)
        self.refresh_name = refresh_name.strip() or DEFAULT_REFRESH_NAME
        self.refresh_class = refresh_class.strip() or DEFAULT_REFRESH_CLASS
        self.multithread = multithread
        self.thread_count = max(1, thread_count if multithread else 1)
        self.task_mode = task_mode
        self.task_name = "锁定" if task_mode == "lock" else "查询"
        self.stop_event = threading.Event()
        self.pause_event = threading.Event()
        self.done = 0
        self.success = 0
        self.failed = 0
        self.failures: list[QueryFailure] = []
        self.started_at = 0.0
        self.stats_lock = threading.Lock()
        self.failure_lock = threading.Lock()
        self.clients_lock = threading.Lock()
        self.clients: dict[int, dict[str, Any]] = {}
        self.writer = make_result_writer(output_path)
        self.solver = CaptchaSolver(paths)
        self.thread: threading.Thread | None = None
        self.run_id = datetime.now().strftime("%Y%m%d%H%M%S%f")

    def start(self) -> None:
        self.thread = threading.Thread(target=self._run, name="YiChaFenQueryRunner", daemon=True)
        self.thread.start()

    def stop(self) -> None:
        self.stop_event.set()
        self.pause_event.clear()

    def pause(self) -> None:
        self.pause_event.set()

    def resume(self) -> None:
        self.pause_event.clear()

    def is_paused(self) -> bool:
        return self.pause_event.is_set()

    def emit(self, kind: str, **payload: Any) -> None:
        payload["kind"] = kind
        payload["run_id"] = self.run_id
        self.events.put(payload)

    def log(self, message: str) -> None:
        self.emit("log", message=message)

    def _run(self) -> None:
        self.started_at = time.monotonic()
        self.log(
            f"开始{self.task_name}：{self.query_info.name}，总计 {len(self.rows)} 条，"
            f"线程数 {self.thread_count}，验证码策略 {self.strategy}"
        )
        self._emit_progress()
        try:
            if self.multithread:
                with ThreadPoolExecutor(max_workers=self.thread_count) as executor:
                    futures = [
                        executor.submit(self._process_row, index, row)
                        for index, row in enumerate(self.rows, start=1)
                    ]
                    for future in as_completed(futures):
                        if self.stop_event.is_set():
                            for item in futures:
                                item.cancel()
                        try:
                            future.result()
                        except Exception as exc:
                            self.log(f"后台任务异常：{exc}")
            else:
                for index, row in enumerate(self.rows, start=1):
                    if self.stop_event.is_set():
                        break
                    self._wait_if_paused()
                    self._process_row(index, row)
        except Exception:
            self.emit("error", message=traceback.format_exc())
        finally:
            status = "已停止" if self.stop_event.is_set() else "已完成"
            self.log(
                f"{self.task_name}{status}：成功 {self.success} 条，失败 {self.failed} 条，"
                f"输出文件 {self.output_path}"
            )
            self.emit(
                "finished",
                stopped=self.stop_event.is_set(),
                success=self.success,
                failed=self.failed,
                output=str(self.output_path),
                failures=[
                    {
                        "time": item.time,
                        "index": item.index,
                        "data": item.data,
                        "reason": item.reason,
                    }
                    for item in self._failure_snapshot()
                ],
            )

    def _get_context(self) -> dict[str, Any]:
        ident = threading.get_ident()
        with self.clients_lock:
            if ident not in self.clients:
                client = YichafenClient(self.paths, f"worker_{ident}")
                client.open_query_page(self.query_info.url)
                self.clients[ident] = {"client": client, "since_refresh": 0}
                self.log(f"线程 {ident} 已初始化会话 Cookie")
            return self.clients[ident]

    def _get_lock_context(self, index: int) -> dict[str, Any]:
        ident = threading.get_ident()
        cookie_name = f"lock_{self.run_id}_{index}_{ident}_{int(time.time() * 1000)}"
        client = YichafenClient(self.paths, cookie_name)
        client.open_query_page(self.query_info.url)
        self.log(f"[{index}] 已初始化独立锁定会话")
        return {"client": client, "since_refresh": 0}

    def _process_row(self, index: int, row: dict[str, str]) -> None:
        self._wait_if_paused()
        if self.stop_event.is_set():
            return

        params = self._build_params(row)
        label = self._params_label(params)
        self._wait_if_paused()
        if self.stop_event.is_set():
            return
        context = self._get_lock_context(index) if self.task_mode == "lock" else self._get_context()
        client: YichafenClient = context["client"]

        try:
            attempt = self._query_with_strategy(client, params, allow_ocr=self.strategy in {"ocr", "mixed"})
            for retry_index in range(1, 3):
                if not self._needs_fresh_lock_session(attempt.message):
                    break
                self.log(f"[{index}] 当前会话已锁定过其他查询，正在更换独立会话重试 {retry_index}/2")
                context = self._get_lock_context(index)
                client = context["client"]
                attempt = self._query_with_strategy(client, params, allow_ocr=self.strategy in {"ocr", "mixed"})
            if attempt.data and (attempt.success or self.task_mode == "lock"):
                self.writer.append(attempt.data)
            if attempt.success:
                self._mark_done(success=True)
                ocr_suffix = "，OCR" if attempt.used_ocr else ""
                self.log(f"[{index}] 成功{ocr_suffix}：{label}")
            else:
                if self.task_mode == "lock" and not attempt.data:
                    self.writer.append(self._build_lock_output(params, success=False))
                self._mark_done(success=False, index=index, data=label, reason=attempt.message)
                self.log(f"[{index}] 失败：{label}；{attempt.message}")
        except Exception as exc:
            self._mark_done(success=False, index=index, data=label, reason=str(exc))
            self.log(f"[{index}] 异常：{label}；{exc}")

        if self.task_mode != "lock" and self.strategy in {"refresh", "mixed"} and not self.stop_event.is_set():
            self._wait_if_paused()
            if self.stop_event.is_set():
                return
            context["since_refresh"] += 1
            if context["since_refresh"] >= 2:
                self._send_refresh_request(client, context)

    def _build_params(self, row: dict[str, str]) -> dict[str, str]:
        params: dict[str, str] = {}
        for field in self.fields:
            csv_column = self.mappings[field.name]
            params[field.name] = clean_text(row.get(csv_column, ""))
        return params

    def _build_refresh_params(self) -> dict[str, str]:
        params: dict[str, str] = {}
        for field in self.fields:
            probe = f"{field.label} {field.name}".lower()
            if "姓名" in probe or "xingming" in probe or "name" in probe:
                params[field.name] = self.refresh_name
            elif "班级" in probe or "banji" in probe or "class" in probe:
                params[field.name] = self.refresh_class
            else:
                params[field.name] = ""
        return params

    def _params_label(self, params: dict[str, str]) -> str:
        parts = []
        for field in self.fields:
            parts.append(f"{field.label}={params.get(field.name, '')}")
        return "，".join(parts)

    def _build_lock_output(self, params: dict[str, str], *, success: bool) -> dict[str, str]:
        data = {field.label: params.get(field.name, "") for field in self.fields}
        data["锁定是否成功"] = "是" if success else "否"
        return data

    @staticmethod
    def _is_already_locked_message(message: str) -> bool:
        return "查询已锁定" in message or "已被锁定" in message

    @staticmethod
    def _needs_fresh_lock_session(message: str) -> bool:
        return "已经锁定过一个查询" in message or "不要锁定别人" in message

    def _query_with_strategy(
        self,
        client: YichafenClient,
        params: dict[str, str],
        *,
        allow_ocr: bool,
    ) -> QueryAttempt:
        self._wait_if_paused()
        if self.stop_event.is_set():
            return QueryAttempt(False, "用户停止")
        response = client.submit_conditions(self.post_url, params)
        return self._handle_submit_response(client, params, response, allow_ocr=allow_ocr)

    def _handle_submit_response(
        self,
        client: YichafenClient,
        params: dict[str, str],
        response: dict[str, Any],
        *,
        allow_ocr: bool,
    ) -> QueryAttempt:
        if int(response.get("status", 0)) == 1:
            return self._handle_successful_submit(client, params, response)

        info = clean_text(response.get("info", "查询失败"))
        if self.task_mode == "lock" and self._is_already_locked_message(info):
            return QueryAttempt(True, info, data=self._build_lock_output(params, success=True))
        captcha_required = bool(response.get("showPicVerify")) or "验证码" in info
        if captcha_required and allow_ocr:
            self.log(f"检测到验证码，进入 OCR 重试：{info}")
            return self._retry_with_ocr(client, params)
        return QueryAttempt(False, info, captcha_required=captcha_required)

    def _handle_successful_submit(
        self,
        client: YichafenClient,
        params: dict[str, str],
        response: dict[str, Any],
        *,
        used_ocr: bool = False,
    ) -> QueryAttempt:
        result_url = response.get("url", "")
        if not result_url:
            return QueryAttempt(False, "提交成功但未返回结果页 URL")
        self._wait_if_paused()
        if self.stop_event.is_set():
            return QueryAttempt(False, "用户停止")

        html = client.get_result_page(result_url)
        if self.task_mode == "lock":
            result_page_url = urljoin(client.base_url + "/", result_url)
            try:
                lock_url, success_url = YichafenParser.parse_lock_request(result_page_url, html)
                self._wait_if_paused()
                if self.stop_event.is_set():
                    return QueryAttempt(False, "用户停止")
                lock_result = client.submit_lock(lock_url, referer=result_page_url)
                success = int(lock_result.get("status", 0)) == 1
                info = clean_text(lock_result.get("info", ""))
                if success and success_url:
                    message = f"锁定成功：{success_url}"
                elif success:
                    message = "锁定成功"
                else:
                    message = info or "锁定接口返回失败"
                if not success and self._is_already_locked_message(message):
                    success = True
                return QueryAttempt(
                    success,
                    message,
                    data=self._build_lock_output(params, success=success),
                    used_ocr=used_ocr,
                )
            except Exception as exc:
                return QueryAttempt(
                    False,
                    f"锁定失败：{exc}",
                    data=self._build_lock_output(params, success=False),
                    used_ocr=used_ocr,
                )

        data = YichafenParser.parse_result(html)
        return QueryAttempt(True, response.get("info", "查询成功"), data=data, used_ocr=used_ocr)

    def _retry_with_ocr(self, client: YichafenClient, params: dict[str, str]) -> QueryAttempt:
        last_message = "验证码识别失败"
        for attempt_index in range(1, self.ocr_retries + 1):
            self._wait_if_paused()
            if self.stop_event.is_set():
                return QueryAttempt(False, "用户停止")
            image_bytes = client.get_captcha_image()
            image_file = self.paths.ocr / f"captcha_{int(time.time() * 1000)}_{attempt_index}.png"
            image_file.write_bytes(image_bytes)
            try:
                code = self.solver.solve(image_bytes)
            except Exception as exc:
                return QueryAttempt(False, f"OCR 初始化或识别失败：{exc}", captcha_required=True)
            self.log(f"OCR 第 {attempt_index}/{self.ocr_retries} 次识别结果：{code or '<空>'}")
            if not code:
                last_message = "OCR 返回空结果"
                continue

            retry_params = dict(params)
            retry_params["verify"] = code
            self._wait_if_paused()
            if self.stop_event.is_set():
                return QueryAttempt(False, "用户停止")
            response = client.submit_conditions(self.post_url, retry_params)
            if int(response.get("status", 0)) == 1:
                return self._handle_successful_submit(client, params, response, used_ocr=True)

            last_message = clean_text(response.get("info", "验证码重试失败"))
            captcha_required = bool(response.get("showPicVerify")) or "验证码" in last_message
            if not captcha_required:
                return QueryAttempt(False, last_message)

        return QueryAttempt(False, f"{last_message}，已达到 OCR 重试上限", captcha_required=True)

    def _send_refresh_request(self, client: YichafenClient, context: dict[str, Any]) -> None:
        self._wait_if_paused()
        if self.stop_event.is_set():
            return
        params = self._build_refresh_params()
        label = self._params_label(params)
        self.log(f"发送刷新请求：{label}")
        try:
            attempt = self._query_with_strategy(
                client,
                params,
                allow_ocr=self.strategy == "mixed",
            )
            if attempt.success:
                context["since_refresh"] = 0
                suffix = "，已用 OCR 兜底" if attempt.used_ocr else ""
                self.log(f"刷新请求成功{suffix}")
            else:
                self.log(f"刷新请求失败：{attempt.message}")
        except Exception as exc:
            self.log(f"刷新请求异常：{exc}")

    def _mark_done(
        self,
        *,
        success: bool,
        index: int | None = None,
        data: str = "",
        reason: str = "",
    ) -> None:
        with self.stats_lock:
            self.done += 1
            if success:
                self.success += 1
            else:
                self.failed += 1
        if not success and index is not None:
            self._record_failure(index=index, data=data, reason=reason)
        self._emit_progress()

    def _record_failure(self, *, index: int, data: str, reason: str) -> None:
        failure = QueryFailure(
            time=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            index=index,
            data=data or "空数据",
            reason=reason or "未知原因",
        )
        with self.failure_lock:
            self.failures.append(failure)

    def _failure_snapshot(self) -> list[QueryFailure]:
        with self.failure_lock:
            return list(self.failures)

    def _wait_if_paused(self) -> None:
        while self.pause_event.is_set() and not self.stop_event.is_set():
            time.sleep(0.1)

    def _emit_progress(self) -> None:
        elapsed = max(0.0, time.monotonic() - self.started_at) if self.started_at else 0.0
        speed = self.done / elapsed if elapsed > 0 else 0.0
        remaining = max(0, len(self.rows) - self.done)
        eta = remaining / speed if speed > 0 else 0.0
        self.emit(
            "progress",
            done=self.done,
            total=len(self.rows),
            success=self.success,
            failed=self.failed,
            elapsed=elapsed,
            eta=eta,
            speed=speed,
        )
